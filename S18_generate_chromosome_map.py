"""
S18_generate_chromosome_map.py

Generate chromosome map SVG for recurrent candidate genes.

Maps recurrent candidate genes onto the B. subtilis 168
chromosome (AL009126.3), highlighting the SPβ prophage
genes.

Author: Eva Stare
"""

import re
try:
    import openpyxl
except ImportError:
    raise ImportError("Please install openpyxl: pip install openpyxl")


# =============================================================================
# 1. LOAD AND PARSE DATA
# =============================================================================

REVIEWED_FILE = "strain_centric_recurrent_candidates_annotated_reviewed.xlsx"
GENOME_SIZE = 4_215_606  # B. subtilis 168 (AL009126.3)

# SVG coordinate mapping
SVG_X_START = 40
SVG_X_END = 640
SVG_WIDTH = SVG_X_END - SVG_X_START  # 600 px


def bp_to_x(bp):
    """Convert a base-pair position to SVG x coordinate."""
    return SVG_X_START + (bp / GENOME_SIZE) * SVG_WIDTH


def parse_location(loc_str):
    """Extract start and end positions from location string.
    Handles formats like '2,281,667 → 2,282,485' or '2,275,200 ← 2,275,697'
    """
    if not loc_str:
        return None, None
    loc = loc_str.replace("\xa0", " ").replace(",", "")
    nums = re.findall(r"(\d+)", loc)
    if len(nums) >= 2:
        return int(nums[0]), int(nums[1])
    return None, None


# Load reviewed candidates
wb = openpyxl.load_workbook(REVIEWED_FILE)
ws = wb.active

genes = []
for r in range(2, ws.max_row + 1):
    gene = ws.cell(r, 1).value
    loc = ws.cell(r, 4).value or ""
    broad = ws.cell(r, 3).value or ""
    if not gene:
        continue
    # Skip non-gene features and hypothetical proteins
    gene_str = str(gene)
    if (gene_str.startswith("hypothetical protein") or
        gene_str.startswith("BF") or
        gene_str.startswith("PHAGE_") or
        gene_str.startswith("Novel_") or
        gene_str == "sublancin_168"):
        continue

    start, end = parse_location(loc)
    if start and end:
        is_spbeta = "sp-beta" in broad.lower() or "spbeta" in broad.lower()
        genes.append({
            "name": gene_str,
            "start": min(start, end),
            "end": max(start, end),
            "mid": (start + end) / 2,
            "broad": broad,
            "spbeta": is_spbeta,
        })

genes.sort(key=lambda x: x["start"])

n_spbeta = sum(1 for g in genes if g["spbeta"])
n_nonspbeta = sum(1 for g in genes if not g["spbeta"])
n_total = len(genes)

# SPβ region boundaries
spb_genes = [g for g in genes if g["spbeta"]]
spb_start = min(g["start"] for g in spb_genes)
spb_end = max(g["end"] for g in spb_genes)
spb_size_kb = (spb_end - spb_start) / 1000

# Non-SPβ genes
non_spb = [g for g in genes if not g["spbeta"]]

print(f"Named genes with coordinates: {n_total}")
print(f"  SPβ: {n_spbeta}")
print(f"  Non-SPβ: {n_nonspbeta}")
print(f"SPβ region: {spb_start:,} - {spb_end:,} ({spb_size_kb:.1f} kb)")
print(f"\nNon-SPβ genes:")
for g in non_spb:
    print(f"  {g['name']:<40} {g['start']:>10,} - {g['end']:>10,}  | {g['broad'][:50]}")


# =============================================================================
# 2. DEFINE LABEL ANNOTATIONS FOR NON-SPβ GENES
# =============================================================================

# Manual label assignments for clarity on the map
# (gene_name_prefix, label_text, sublabel, color, label_y_offset, label_side)
label_annotations = {
    "conE": {"label": "conE", "sub": "ICEBs1", "color": "#D85A30", "y": 212},
    "rapA": {"label": "rapA", "sub": "Quorum sensing", "color": "#1D9E75", "y": 252},
    "xerC": {"label": "xerC/codV", "sub": "Integrase", "color": "#D85A30", "y": 212},
    "yneA": {"label": "yneA", "sub": None, "color": "#888780", "y": 342, "rotate": True},
    "yoeC": {"label": "yoeC", "sub": None, "color": "#888780", "y": 342, "rotate": True},
    "yozM": {"label": "yozM", "sub": None, "color": "#888780", "y": 342, "rotate": True},
    "dinG": {"label": "dinG", "sub": "DNA repair", "color": "#D85A30", "y": 212},
    "yqbF": {"label": "yqbF", "sub": "skin element", "color": "#BA7517", "y": 252},
    "tdk": {"label": "tdk", "sub": None, "color": "#888780", "y": 212},
    "rapF": {"label": "rapF", "sub": "Quorum sensing", "color": "#1D9E75", "y": 212},
}

# SPβ flanking genes (special handling)
spb_flank_left = ["cgeA", "spsMn"]
spb_flank_right = ["spsMc"]


# =============================================================================
# 3. BUILD SVG
# =============================================================================

def make_scale_tick(mb_value):
    """Generate SVG for a scale tick at a given Mb position."""
    x = bp_to_x(mb_value * 1_000_000)
    return (
        f'  <line x1="{x:.0f}" y1="140" x2="{x:.0f}" y2="148" stroke="#5F5E5A" stroke-width="0.5"/>\n'
        f'  <text class="ts" x="{x:.0f}" y="160" text-anchor="middle">{mb_value:.1f}</text>\n'
    )


def make_gene_tick(gene, y_top=106, y_bot=120):
    """Generate SVG tick mark for a gene on the chromosome."""
    x = bp_to_x(gene["mid"])
    color = "#7F77DD" if gene["spbeta"] else "#888780"
    width = 2 if not gene["spbeta"] else 1.5
    return f'  <line x1="{x:.0f}" y1="{y_top}" x2="{x:.0f}" y2="{y_bot}" stroke="{color}" stroke-width="{width}"/>\n'


# Compute SVG positions
spb_x_start = bp_to_x(spb_start)
spb_x_end = bp_to_x(spb_end)
spb_x_width = spb_x_end - spb_x_start

svg_parts = []

# Header
svg_parts.append(
    '<svg width="680" height="520" viewBox="0 0 680 520" xmlns="http://www.w3.org/2000/svg">\n'
    "  <style>\n"
    "    text { font-family: 'Helvetica Neue', Arial, sans-serif; }\n"
    "    .th { font-size: 14px; font-weight: 500; fill: #2C2C2A; }\n"
    "    .ts { font-size: 12px; font-weight: 400; fill: #5F5E5A; }\n"
    "    .ts10 { font-size: 10px; font-weight: 400; }\n"
    "    .ts9 { font-size: 9px; font-weight: 400; fill: #888780; }\n"
    "  </style>\n\n"
)

# Title
svg_parts.append(
    f'  <text class="th" x="340" y="24" text-anchor="middle">'
    f"Chromosomal distribution of {n_total} named recurrent candidate genes</text>\n"
    f'  <text class="ts" x="340" y="42" text-anchor="middle">'
    f'Mapped onto the <tspan style="font-style:italic">B. subtilis</tspan>'
    f" 168 chromosome ({GENOME_SIZE/1e6:.2f} Mb, AL009126.3; Subtiwiki v5)</text>\n\n"
)

# Chromosome backbone
svg_parts.append(
    '  <rect x="40" y="120" width="600" height="14" rx="7" '
    'fill="#F1EFE8" stroke="#B4B2A9" stroke-width="0.5"/>\n\n'
)

# Scale ticks (every 0.5 Mb)
for mb in [0, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0]:
    svg_parts.append(make_scale_tick(mb))
svg_parts.append('  <text class="ts" x="640" y="160" text-anchor="end">Mb</text>\n\n')

# SPβ region highlight
svg_parts.append(
    f'  <!-- SPβ prophage region: {spb_start:,} - {spb_end:,} -->\n'
    f'  <rect x="{spb_x_start:.0f}" y="68" width="{spb_x_width:.0f}" height="66" '
    f'rx="2" fill="#7F77DD" opacity="0.12"/>\n'
    f'  <rect x="{spb_x_start:.0f}" y="120" width="{spb_x_width:.0f}" height="14" '
    f'rx="0" fill="#7F77DD" opacity="0.35"/>\n'
    f'  <text class="ts" x="{(spb_x_start + spb_x_end) / 2:.0f}" y="64" '
    f'text-anchor="middle" fill="#534AB7" font-weight="500">SPβ prophage</text>\n'
    f'  <text class="ts9" x="{(spb_x_start + spb_x_end) / 2:.0f}" y="80" '
    f'text-anchor="middle" fill="#534AB7">{spb_size_kb:.0f} kb</text>\n'
    f'  <text class="ts9" x="{(spb_x_start + spb_x_end) / 2:.0f}" y="92" '
    f'text-anchor="middle" fill="#534AB7">{n_spbeta} genes</text>\n\n'
)

# Dense SPβ gene block
svg_parts.append(
    f'  <rect x="{spb_x_start:.0f}" y="107" width="{spb_x_width:.0f}" height="10" '
    f'fill="#7F77DD" opacity="0.55" rx="1"/>\n\n'
)

# Non-SPβ gene ticks with labels
for gene in non_spb:
    x = bp_to_x(gene["mid"])
    name = gene["name"]

    # Find matching annotation
    ann = None
    for prefix, a in label_annotations.items():
        if name.startswith(prefix):
            ann = a
            break

    # Check if SPβ flanking
    is_flank = any(name.startswith(p) for p in spb_flank_left + spb_flank_right)
    color = "#639922" if is_flank else (ann["color"] if ann else "#888780")
    tick_width = 1.5 if color == "#888780" else 2

    # Tick mark
    svg_parts.append(
        f'  <line x1="{x:.0f}" y1="106" x2="{x:.0f}" y2="120" '
        f'stroke="{color}" stroke-width="{tick_width}"/>\n'
    )

    # Label with leader line
    if ann:
        y_label = ann["y"]
        rotate = ann.get("rotate", False)

        # Leader line
        if rotate:
            svg_parts.append(
                f'  <line x1="{x:.0f}" y1="170" x2="{x:.0f}" y2="330" '
                f'stroke="{color}" stroke-width="0.5" stroke-dasharray="2 2"/>\n'
            )
            svg_parts.append(
                f'  <text class="ts9" x="{x:.0f}" y="{y_label}" text-anchor="end" '
                f'transform="rotate(-45, {x:.0f}, {y_label})">{ann["label"]}</text>\n'
            )
        else:
            # Offset tdk left and rapF right to avoid overlap
            if ann["label"] == "tdk":
                label_x = x - 10
            elif ann["label"] == "rapF":
                label_x = x + 13
            else:
                label_x = x
            # Italic gene name using tspan
            italic_label = f'<tspan style="font-style:italic">{ann["label"]}</tspan>'
            svg_parts.append(
                f'  <line x1="{x:.0f}" y1="170" x2="{label_x:.0f}" y2="{y_label - 12}" '
                f'stroke="{color}" stroke-width="0.5" stroke-dasharray="2 2"/>\n'
            )
            svg_parts.append(
                f'  <text class="ts10" x="{label_x:.0f}" y="{y_label}" '
                f'text-anchor="middle" fill="{color}">{italic_label}</text>\n'
            )
            if ann.get("sub"):
                # Offset sublabel x for rapF to avoid overlap with tdk
                sub_x = label_x + 16 if ann["label"] == "rapF" else label_x
                svg_parts.append(
                    f'  <text class="ts9" x="{sub_x:.0f}" y="{y_label + 12}" '
                    f'text-anchor="middle">{ann["sub"]}</text>\n'
                )

    elif is_flank:
        # SPβ flanking genes get special grouped labels
        pass  # Handled below as grouped annotations

# SPβ flanking gene labels (grouped)
left_flank_x = bp_to_x(non_spb[0]["mid"]) if non_spb else spb_x_start
# Find cgeA and spsMn positions
cgeA_x = None
spsM_left_x = None
spsM_right_x = None
for g in non_spb:
    if g["name"].startswith("cgeA"):
        cgeA_x = bp_to_x(g["mid"])
    elif g["name"].startswith("spsMn"):
        spsM_left_x = bp_to_x(g["mid"])
    elif g["name"].startswith("spsMc"):
        spsM_right_x = bp_to_x(g["mid"])

if cgeA_x and spsM_left_x:
    avg_left = (cgeA_x + spsM_left_x) / 2
    svg_parts.append(
        f'  <line x1="{avg_left:.0f}" y1="170" x2="{avg_left - 40:.0f}" y2="290" '
        f'stroke="#639922" stroke-width="0.5" stroke-dasharray="2 2"/>\n'
        f'  <text class="ts10" x="{avg_left - 40:.0f}" y="302" '
        f'text-anchor="middle" fill="#639922">'
        f'<tspan style="font-style:italic">cgeA</tspan>, '
        f'<tspan style="font-style:italic">spsM</tspan>/'
        f'<tspan style="font-style:italic">yodU</tspan></text>\n'
        f'  <text class="ts9" x="{avg_left - 40:.0f}" y="314" '
        f'text-anchor="middle">SPβ left flank</text>\n'
    )

if spsM_right_x:
    svg_parts.append(
        f'  <line x1="{spsM_right_x:.0f}" y1="170" x2="{spsM_right_x + 40:.0f}" y2="290" '
        f'stroke="#639922" stroke-width="0.5" stroke-dasharray="2 2"/>\n'
        f'  <text class="ts10" x="{spsM_right_x + 40:.0f}" y="302" '
        f'text-anchor="middle" fill="#639922">'
        f'<tspan style="font-style:italic">spsM</tspan>/'
        f'<tspan style="font-style:italic">ypqP</tspan></text>\n'
        f'  <text class="ts9" x="{spsM_right_x + 40:.0f}" y="314" '
        f'text-anchor="middle">SPβ right flank</text>\n'
    )

# SPβ gene count label below
svg_parts.append(
    f'  <text class="ts" x="{(spb_x_start + spb_x_end) / 2:.0f}" y="180" '
    f'text-anchor="middle" fill="#534AB7">{n_spbeta} SPβ genes</text>\n\n'
)

# Legend box
svg_parts.append(
    '  <rect x="40" y="371" width="600" height="135" rx="8" '
    'fill="#F9F8F6" stroke="#B4B2A9" stroke-width="0.5"/>\n\n'
    '  <text class="th" x="60" y="394">Gene distribution summary</text>\n\n'
    '  <rect x="60" y="406" width="12" height="12" rx="2" fill="#7F77DD" opacity="0.55"/>\n'
    f'  <text class="ts" x="80" y="417">SPβ prophage: {n_spbeta} genes in a '
    f'{spb_size_kb:.0f} kb block ({spb_start / 1e6:.2f}\u2013{spb_end / 1e6:.2f} Mb)</text>\n\n'
    '  <rect x="60" y="426" width="12" height="12" rx="2" fill="#639922"/>\n'
    '  <text class="ts" x="80" y="437">SPβ flanking: '
    '<tspan style="font-style:italic">cgeA</tspan> (spore coat) + '
    '<tspan style="font-style:italic">spsM</tspan> variants (spore crust)</text>\n\n'
    '  <rect x="60" y="446" width="12" height="12" rx="2" fill="#1D9E75"/>\n'
    '  <text class="ts" x="80" y="457">Quorum sensing: '
    '<tspan style="font-style:italic">rapA</tspan> (1.32 Mb), '
    '<tspan style="font-style:italic">rapF</tspan> (3.85 Mb)</text>\n\n'
    '  <rect x="60" y="466" width="12" height="12" rx="2" fill="#D85A30"/>\n'
    '  <text class="ts" x="80" y="477">Other: '
    '<tspan style="font-style:italic">conE</tspan>/ICEBs1 (0.54 Mb), '
    '<tspan style="font-style:italic">xerC</tspan> (1.69 Mb), '
    '<tspan style="font-style:italic">dinG</tspan> (2.35 Mb)</text>\n\n'
    '  <rect x="60" y="486" width="12" height="12" rx="2" fill="#888780"/>\n'
    '  <text class="ts" x="80" y="497">Minor/unknown: '
    '<tspan style="font-style:italic">yneA</tspan> (1.92 Mb), '
    '<tspan style="font-style:italic">yoeC</tspan> (2.00 Mb), '
    '<tspan style="font-style:italic">yozM</tspan> (2.07 Mb), '
    '<tspan style="font-style:italic">tdk</tspan> (3.80 Mb)</text>\n\n'
)

# Close SVG
svg_parts.append("</svg>\n")

svg_content = "".join(svg_parts)

# Write output
output_file = "recurrent_candidates_chromosome_map.svg"
with open(output_file, "w", encoding="utf-8") as f:
    f.write(svg_content)

print(f"\nSaved: {output_file}")
print(f"\nSummary:")
print(f"  Total named genes mapped: {n_total}")
print(f"  SPβ cluster: {n_spbeta} genes ({n_spbeta / n_total * 100:.1f}%)")
print(f"  SPβ region: {spb_start:,}\u2013{spb_end:,} bp ({spb_size_kb:.0f} kb)")
print(f"  Non-SPβ genes: {n_nonspbeta} (scattered across {GENOME_SIZE / 1e6:.2f} Mb)")
