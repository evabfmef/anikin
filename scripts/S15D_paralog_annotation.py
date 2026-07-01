"""
Paralog Annotation of Corrected Features

Author: Eva Stare

This script annotates the validated group-specific features with paralog and
split-gene information taken from the manual review of the S15C paralog-check
results.

Pipeline order
--------------
S15B (variant validation)  ->  S15C (paralog check)  ->  S15D (this script).
S15D consumes two MANUALLY CURATED inputs produced after S15C:
  - 67_corrected_features_analysis_reviewed.xlsx
  - 67_paralog_check_results_reviewed.xlsx
Both reviewed files are provided in the supplementary data deposit; the script
cannot be re-run without them.

"""

import pandas as pd
import re, os
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

# Narrow suppression: silence only the pandas/openpyxl UserWarning chatter
# rather than blanket-ignoring every warning, so genuine issues stay visible
# on re-runs.
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
print('Ready!')

BASE_DIR = 'input/'  # UPDATE THIS PATH

CORRECTED_PATH  = os.path.join(BASE_DIR, '04_paralog_input_data/67_corrected_features_analysis_reviewed.xlsx')   # OUTPUT from 02 script (v6)
PARALOG_PATH    = os.path.join(BASE_DIR, '03_paralog_check_output_v5/67_paralog_check_results_reviewed.xlsx')  # OUTPUT from 03 script (manually reviewed!)
MEGAMATRIX_PATH = os.path.join(BASE_DIR, '04_paralog_input_data/67_mega_matrix_full.xlsx')
OUTPUT_DIR      = os.path.join(BASE_DIR, '04_paralog_check_output')
OUTPUT_PATH     = os.path.join(OUTPUT_DIR, '67_corrected_features_annotated_paralogs.xlsx')

os.makedirs(OUTPUT_DIR, exist_ok=True)
print(f'Corrected features: {os.path.exists(CORRECTED_PATH)}')
print(f'Paralog results:    {os.path.exists(PARALOG_PATH)}')
print(f'Megamatrix:         {os.path.exists(MEGAMATRIX_PATH)}')

"""
## 1. Helpers & config
"""

EXCLUDED_STRAIN = None

STRAIN_GROUPS = {
    'g3':    ['RO-F-3'],
    'g4':    ['RS-D-2'],
    'g5':    ['RO-DD-2'],
    'g7':    ['RO-A-4'],
    'g19':   ['NRS6085'],
    'g25_A': ['PS-13', 'PS-14', 'PS-18', 'PS-30', 'PS-31', 'PS-51', 'PS-65',
              'PS-68', 'PS-96', 'PS-168', 'PS-210', 'PS-216', 'PS-233', 'PS-237',
              'MB8_B1', 'MB8_B7', 'MB9_B1', 'MB9_B6', 'P8_B1', 'P9_B1', 'NCIB_3610'],
    'g29':   ['BS16045'],
    'g30':   ['NRS6105', 'NRS6145'],
    'g31':   ['PS-52', 'PS-53'],
    'g34':   ['NRS6128'],
    'g39':   ['PS-209'],
    'g44':   ['PS-196'],
    'g50':   ['PS-108', 'PS-109', 'PS-119', 'PS-130', 'PS-131'],
    'g52':   ['NRS6160'],
    'g53':   ['PS-20', 'PS-24', 'PS-25'],
    'g54':   ['PS-160'],
    'g55':   ['PS-93', 'PS-95'],
    'g56':   ['PS-217', 'PS-218'],
    'g57':   ['PS-15'],
    'g60_A': ['PS-64', 'PS-194', 'NRS6186', 'NRS6181'],
    'g60_B': ['NRS6132', 'NRS6099', 'NRS6187'],
    'g63_A': ['PS-54', 'PS-55', 'PS-149', 'NRS6103', 'NRS6118', 'NRS6127', 'NRS6153'],
    'g82':   ['RO-FF-1'],
    'g92':   ['KF24'],
    'g95':   ['PS-263'],
    'g96':   ['73'],
}

VARPAT = re.compile(r'^(.+)\.(\d+)$')

def get_base_names(gene_str):
    return {(m.group(1) if (m := VARPAT.match(p.strip())) else p.strip())
            for p in str(gene_str).split(',')}

def get_full_suffix(gene_str):
    for p in reversed(str(gene_str).split(',')):
        m = VARPAT.match(p.strip())
        if m: return m.group(2)
    return None

print('Helpers defined.')
print(f'Total groups: {len(STRAIN_GROUPS)}')
print(f'Total strains in groups: {sum(len(v) for v in STRAIN_GROUPS.values())}')

"""
## 2. Load megamatrix
"""

mega = pd.read_excel(MEGAMATRIX_PATH, sheet_name='Sheet1')
strain_cols = [c for c in mega.columns if c != 'Feature_Name']

# In the 67-strain dataset there is no PS-11 to exclude
if EXCLUDED_STRAIN and EXCLUDED_STRAIN in strain_cols:
    strain_cols_filtered = [c for c in strain_cols if c != EXCLUDED_STRAIN]
    print(f'Excluded strain: {EXCLUDED_STRAIN}')
else:
    strain_cols_filtered = strain_cols
    print('No strains excluded.')

N_STRAINS = len(strain_cols_filtered)
print(f'{mega.shape[0]} features, {N_STRAINS} strains')

mega_gene_index = {}
for idx, row in mega.iterrows():
    fname = str(row['Feature_Name'])
    for b in get_base_names(fname):
        mega_gene_index.setdefault(b, []).append((idx, fname, get_full_suffix(fname)))
print(f'Gene index: {len(mega_gene_index)} base names')

"""
## 3. Load paralog results (with manual checks)
"""

par_group_features = pd.read_excel(PARALOG_PATH, sheet_name='GROUP - features summary')
par_group_details  = pd.read_excel(PARALOG_PATH, sheet_name='GROUP - all details')
par_splits         = pd.read_excel(PARALOG_PATH, sheet_name='GROUP - likely splits')
par_all_details    = pd.read_excel(PARALOG_PATH, sheet_name='ALL - all details')

try:
    par_uniqueness = pd.read_excel(PARALOG_PATH, sheet_name='SPLIT uniqueness check')
except (ValueError, KeyError):
    # Sheet is optional; absent in some paralog-result workbooks.
    par_uniqueness = pd.DataFrame()

print(f'GROUP features: {len(par_group_features)}, details: {len(par_group_details)}')
print(f'Splits (manual checked): {len(par_splits)}')
print(f'Uniqueness entries: {len(par_uniqueness)}')

# ---- Build paralog lookup ----
paralog_lookup = {}
for _, row in par_group_features.iterrows():
    paralog_lookup[(row['group'], frozenset(get_base_names(row['feature'])))] = row.to_dict()

# ---- Build split verdict lookup from manual checks ----
# manual_check values are free-text verdicts mapped to three categories:
#   confirmed_split ("yes...")      gene is split, one functional copy    -> blue
#   not_split ("no..."/"not split") gene is not split                     -> yellow
#   unchecked (no entry)            not yet reviewed                      -> orange
#
# not_split covers several free-text sub-types, preserved verbatim in the
# manual_check text but not distinguished by the code, e.g.:
#   true duplicates ("two full-length copies", "X copies"),
#   size differences ("one gene Xbp, second Ybp"),
#   duplicate + fragment, or uncertain cases ("two truncated genes?").

split_verdicts = {}  # (group, frozenset_bases) -> {verdict, manual_check, uniqueness}

for feat_name in par_splits['feature'].unique():
    rows = par_splits[par_splits['feature'] == feat_name]
    group = rows.iloc[0]['group']
    mc = str(rows.iloc[0]['manual_check']).strip() if 'manual_check' in rows.columns and pd.notna(rows.iloc[0].get('manual_check')) else ''
    bases = frozenset(get_base_names(feat_name))

    mc_lower = mc.lower().strip()
    if mc_lower.startswith('yes'):
        verdict = 'confirmed_split'
    elif mc_lower.startswith('no') or 'not split' in mc_lower:
        verdict = 'not_split'
    elif mc == '':
        verdict = 'unchecked'
    else:
        # Fallback: if unclear, mark as unchecked for safety
        verdict = 'unchecked'
        print(f'   Ambiguous manual_check for {group}/{feat_name}: "{mc}" → treating as unchecked')

    uniqueness = ''
    if len(par_uniqueness) > 0:
        uq = par_uniqueness[(par_uniqueness['group'] == group) & (par_uniqueness['feature'] == feat_name)]
        if len(uq) > 0:
            uniqueness = str(uq.iloc[0]['uniqueness'])

    split_verdicts[(group, bases)] = dict(
        verdict=verdict, manual_check=mc, uniqueness=uniqueness, feature=feat_name
    )

n_cs = sum(1 for v in split_verdicts.values() if v['verdict'] == 'confirmed_split')
n_ns = sum(1 for v in split_verdicts.values() if v['verdict'] == 'not_split')
n_uc = sum(1 for v in split_verdicts.values() if v['verdict'] == 'unchecked')
print(f'\nSplit verdicts: {n_cs} confirmed split, {n_ns} not split, {n_uc} unchecked')

# Print verdicts for verification
print(f'\n{"─"*80}')
_VERDICT_TAG = {'confirmed_split': '[split]    ',
                'not_split':       '[not split]',
                'unchecked':       '[unchecked]'}
for (grp, bases), info in sorted(split_verdicts.items(), key=lambda x: (x[0][0], x[1]['feature'])):
    tag = _VERDICT_TAG.get(info['verdict'], '[unchecked]')
    print(f'  {tag} {grp:8s} {info["feature"]:35s} -> {info["verdict"]:20s} | {info["manual_check"][:55]}')

"""
## 4. Annotation functions
"""

def find_paralog_match(group, gene_str):
    qb = get_base_names(gene_str)
    for (pg, pb), info in paralog_lookup.items():
        if pg == group and qb & pb: return info
    return None

def find_split_verdict(group, gene_str):
    qb = get_base_names(gene_str)
    for (sg, sb), info in split_verdicts.items():
        if sg == group and qb & sb: return info
    return None

def get_group_copy_info(group, feature_name):
    qb = get_base_names(feature_name)
    gs = set(STRAIN_GROUPS.get(group, []))
    ng = len(gs)
    for fn in par_group_details['feature'].unique():
        if get_base_names(fn) & qb:
            sub = par_group_details[(par_group_details['group']==group) & (par_group_details['feature']==fn)]
            gr = sub[sub['strain'].isin(gs)]
            if len(gr) > 0:
                copies = gr['n_copies'].values
                nw = len(gr)
                if len(set(copies)) == 1:
                    return f'{int(copies[0])} copies in {nw}/{ng} group strains'
                return f'{int(min(copies))}-{int(max(copies))} copies in {nw}/{ng} group strains'
    return 'multi-copy (details unavailable)'

def check_universal_copies(gene_str):
    bases = get_base_names(gene_str)
    seen = set()
    universal = []
    for b in bases:
        for (ridx, rname, rsuffix) in mega_gene_index.get(b, []):
            if ridx not in seen:
                seen.add(ridx)
                vals = mega.iloc[ridx][strain_cols_filtered].values
                if sum(v == 1 for v in vals) == N_STRAINS:
                    universal.append({'feature_name': rname, 'suffix': rsuffix})
    return universal

def build_annotation(group, gene_str):
    """Returns (paralog_text, category).
    category: 'none' | 'split' | 'not_split_duplicate' | 'duplicate' | 'unchecked'"""
    par_info = find_paralog_match(group, gene_str)
    if par_info is None:
        return 'no', 'none'

    si = find_split_verdict(group, gene_str)
    if si is not None:
        if si['verdict'] == 'confirmed_split':
            txt = 'split gene (1 functional copy)'
            uq = si['uniqueness']
            if 'UNIQUE' in uq.upper():
                txt += ' — split unique to group'
            elif 'ALSO OUTSIDE' in uq.upper():
                txt += ' — split also in other strains'
            elif uq:
                txt += f' — {uq[:50]}'
            return txt, 'split'
        elif si['verdict'] == 'not_split':
            ci = get_group_copy_info(group, gene_str)
            return f'YES, {ci} ({si["manual_check"][:50]})', 'not_split_duplicate'
        elif si['verdict'] == 'unchecked':
            ci = get_group_copy_info(group, gene_str)
            return f'LIKELY SPLIT - needs check ({ci})', 'unchecked'

    # Has paralogs but was not in the likely-splits sheet (i.e., distant paralogs, not adjacent)
    ci = get_group_copy_info(group, gene_str)
    return f'YES, {ci}', 'duplicate'

print('Functions defined.')

"""
## 5. Annotate corrected features
"""

wb = load_workbook(CORRECTED_PATH)

hf = Font(bold=True, size=10, name='Arial')
hfill = PatternFill('solid', fgColor='D9E1F2')
ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
cf = Font(size=9, name='Arial')
ca = Alignment(vertical='top', wrap_text=True)
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

yellow  = PatternFill('solid', fgColor='FFFF00')   # true duplicates
green   = PatternFill('solid', fgColor='C6EFCE')   # no universal copy
red     = PatternFill('solid', fgColor='FFC7CE')   # universal copy exists
blue    = PatternFill('solid', fgColor='BDD7EE')   # confirmed split
orange  = PatternFill('solid', fgColor='FCE4D6')   # unchecked

summary_stats = {}

for sn in wb.sheetnames:
    if sn in ('SUMMARY', 'Removed artifacts'): continue
    ws = wb[sn]
    mr = ws.max_row

    ws.insert_cols(7, 2)   # after present status (F)
    ws.insert_cols(16, 2)  # after absent status (shifted)

    for col, lbl in [(7,'paralogs'),(8,'universal copy'),(16,'paralogs'),(17,'universal copy')]:
        c = ws.cell(row=2, column=col, value=lbl)
        c.font=hf; c.fill=hfill; c.alignment=ha; c.border=tb
    ws.column_dimensions['G'].width = 42
    ws.column_dimensions['H'].width = 28
    ws.column_dimensions[get_column_letter(16)].width = 42
    ws.column_dimensions[get_column_letter(17)].width = 28

    st = {'dp':0,'sp':0,'up':0,'da':0,'sa':0,'ua':0}

    for rn in range(3, mr+1):
        # PRESENT (col 1)
        pg = ws.cell(row=rn, column=1).value
        if pg and str(pg).strip():
            pg = str(pg).strip()
            ptxt, pcat = build_annotation(sn, pg)
            c7 = ws.cell(row=rn, column=7, value=ptxt)
            c7.font=cf; c7.alignment=ca; c7.border=tb
            c8 = ws.cell(row=rn, column=8)
            c8.font=cf; c8.alignment=ca; c8.border=tb

            if pcat == 'split':
                c7.fill = blue; c8.value = '—'; st['sp'] += 1
            elif pcat == 'unchecked':
                c7.fill = orange; c8.value = '—'
            elif pcat in ('duplicate', 'not_split_duplicate'):
                c7.fill = yellow; st['dp'] += 1
                uni = check_universal_copies(pg)
                if uni:
                    c8.value = 'YES — universal copy found'; c8.fill = red; st['up'] += 1
                else:
                    c8.value = 'No universal copy'; c8.fill = green
            else:
                c8.value = '—'

        # ABSENT (col 10 after insert)
        ag = ws.cell(row=rn, column=10).value
        if ag and str(ag).strip():
            ag = str(ag).strip()
            atxt, acat = build_annotation(sn, ag)
            c16 = ws.cell(row=rn, column=16, value=atxt)
            c16.font=cf; c16.alignment=ca; c16.border=tb
            c17 = ws.cell(row=rn, column=17)
            c17.font=cf; c17.alignment=ca; c17.border=tb

            if acat == 'split':
                c16.fill = blue; c17.value = '—'; st['sa'] += 1
            elif acat == 'unchecked':
                c16.fill = orange; c17.value = '—'
            elif acat in ('duplicate', 'not_split_duplicate'):
                c16.fill = yellow; st['da'] += 1
                uni = check_universal_copies(ag)
                if uni:
                    c17.value = 'YES — universal copy found'; c17.fill = red; st['ua'] += 1
                else:
                    c17.value = 'No universal copy'; c17.fill = green
            else:
                c16.value = 'no'; c17.value = '—'

    summary_stats[sn] = st
    print(f'  {sn}: pres[{st["dp"]} dup, {st["sp"]} split, {st["up"]} uni] '
          f'abs[{st["da"]} dup, {st["sa"]} split, {st["ua"]} uni]')

"""
## 6. Update SUMMARY & save
"""

if 'SUMMARY' in wb.sheetnames:
    ws = wb['SUMMARY']
    nc = ws.max_column + 2
    for i, lbl in enumerate(['dup_pres','split_pres','uni_pres','dup_abs','split_abs','uni_abs']):
        ws.cell(row=1, column=nc+i, value=lbl).font = hf
    for r in range(2, ws.max_row+1):
        g = str(ws.cell(row=r, column=1).value).strip()
        if g in summary_stats:
            s = summary_stats[g]
            for i, k in enumerate(['dp','sp','up','da','sa','ua']):
                ws.cell(row=r, column=nc+i, value=s[k])

wb.save(OUTPUT_PATH)
print(f'\nSaved: {OUTPUT_PATH}')

"""
## 7. Summary table
"""

print(f'{"Group":8s} | {"PRESENT":^22s} | {"ABSENT":^22s}')
print(f'{"":8s} | {"Dup":>4s} {"Split":>5s} {"Uni":>4s} | {"Dup":>4s} {"Split":>5s} {"Uni":>4s}')
print('-' * 55)
for g in sorted(STRAIN_GROUPS.keys(), key=lambda x: (x.split('_')[0].replace('g','').zfill(3), x)):
    s = summary_stats.get(g, {'dp':0,'sp':0,'up':0,'da':0,'sa':0,'ua':0})
    print(f'{g:8s} | {s["dp"]:>4} {s["sp"]:>5} {s["up"]:>4} | {s["da"]:>4} {s["sa"]:>5} {s["ua"]:>4}')

print('\nCell fill colour legend (in the output Excel):')
print('  Blue   = confirmed split (1 functional copy)')
print('  Yellow = true duplicate (multiple genomic copies)')
print('  Green  = no universal copy (group-specific variant)')
print('  Red    = universal copy exists (common allele present in all strains)')
print('  Orange = likely split, not yet manually checked')
